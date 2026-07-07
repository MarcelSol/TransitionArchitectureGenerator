"""
excel_exporter.py

Exports a TransitionModel to an Excel workbook.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font

from tag.builder import TransitionModelBuilder
from tag.transition_model import (
    TransitionModel,
    NodeCategory,
    InterfaceDirection,
    TransferType,
)

from tag.validation import Validator
from tag.validation_report import ValidationSeverity


class ExcelExporter:

    @staticmethod
    def export(
        model: TransitionModel,
        filename: str,
    ) -> None:

        workbook = Workbook()
        report = Validator.validate(model)

        #
        # Remove the default worksheet created by openpyxl.
        #
        workbook.remove(workbook.active)

        ExcelExporter._write_nodes(
            workbook,
            model,
        )

        ExcelExporter._write_interfaces(
            workbook,
            model,
        )

        ExcelExporter._write_summary(
            workbook,
            model,
        )

        ExcelExporter._write_validation_sheet(
            workbook,
            report,
        )

        workbook.save(filename)

    # ---------------------------------------------------------------------

    @staticmethod
    def _write_nodes(
        workbook: Workbook,
        model: TransitionModel,
    ) -> None:

        sheet = workbook.create_sheet("Nodes")

        milestones = model.milestones

        headers = [
            "ID",
            "Name",
            "Category",
            "First Appears",
            "Disappears",
        ] + milestones

        sheet.append(headers)

        ExcelExporter._format_header(sheet)

        final_milestone = milestones[-1]

        for node in sorted(
            model.nodes.values(),
            key=lambda n: n.name.lower(),
        ):

            first = TransitionModelBuilder._first_visible(
                node.visible_on,
                milestones,
            )

            retired = TransitionModelBuilder._retired_in(
                node.visible_on,
                milestones,
            )

            if retired == final_milestone:
                retired = ""

            row = [
                node.id,
                node.name,
                node.category.value,
                first,
                retired,
            ]

            for milestone in milestones:
                row.append(
                    "X"
                    if milestone in node.visible_on
                    else ""
                )

            sheet.append(row)

        ExcelExporter._auto_width(sheet)

    # ---------------------------------------------------------------------

    @staticmethod
    def _format_header(sheet):

        bold = Font(bold=True)

        for cell in sheet[1]:
            cell.font = bold

        sheet.freeze_panes = "A2"

    # ---------------------------------------------------------------------

    @staticmethod
    def _auto_width(sheet):

        for column in sheet.columns:

            length = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column
            )

            sheet.column_dimensions[
                column[0].column_letter
            ].width = length + 2

    # ---------------------------------------------------------------------

    @staticmethod
    def _write_interfaces(
        workbook: Workbook,
        model: TransitionModel,
    ) -> None:

        sheet = workbook.create_sheet("Interfaces")

        milestones = model.milestones

        headers = [
            "ID",
            "Source",
            "Target",
            "Transfer",
            "Direction",
            "First Appears",
            "Disappears",
        ] + milestones

        sheet.append(headers)

        ExcelExporter._format_header(sheet)

        final_milestone = milestones[-1]

        node_lookup = {
            node.id: node
            for node in model.nodes.values()
        }

        for interface in sorted(
            model.interfaces.values(),
            key=lambda i: (
                node_lookup[i.source].name.lower(),
                node_lookup[i.target].name.lower(),
            ),
        ):

            first = TransitionModelBuilder._first_visible(
                interface.visible_on,
                milestones,
            )

            last = TransitionModelBuilder._last_visible(
                interface.visible_on,
                milestones,
            )

            if last == final_milestone:
                last = ""

            source = node_lookup[interface.source].name
            target = node_lookup[interface.target].name

            transfer = (
                "Manual"
                if interface.transfer_type == TransferType.MANUAL
                else "Automatic"
            )

            direction = (
                "Two-way"
                if interface.direction == InterfaceDirection.TWO_WAY
                else "One-way"
            )

            row = [
                interface.id,
                source,
                target,
                transfer,
                direction,
                first,
                last,
            ]

            for milestone in milestones:

                row.append(
                    "X"
                    if milestone in interface.visible_on
                    else ""
                )

            sheet.append(row)

        ExcelExporter._auto_width(sheet)

    # ---------------------------------------------------------------------

    @staticmethod
    def _write_summary(
        workbook: Workbook,
        model: TransitionModel,
    ) -> None:

        sheet = workbook.create_sheet("Summary")

        headers = [
            "Milestone",
            "New Nodes",
            "Retired Nodes",
            "Active Nodes",
            "New Interfaces",
            "Retired Interfaces",
            "Active Interfaces",
        ]

        sheet.append(headers)

        ExcelExporter._format_header(sheet)

        milestones = model.milestones

        final_milestone = milestones[-1]

        for milestone in milestones:

            new_nodes = 0
            retired_nodes = 0
            active_nodes = 0

            for node in model.nodes.values():

                first = TransitionModelBuilder._first_visible(
                    node.visible_on,
                    milestones,
                )

                retired = TransitionModelBuilder._retired_in(
                    node.visible_on,
                    milestones,
                )

                if milestone in node.visible_on:
                    active_nodes += 1

                if first == milestone:
                    new_nodes += 1

                if retired == milestone:
                    retired_nodes += 1

            new_interfaces = 0
            retired_interfaces = 0
            active_interfaces = 0

            for interface in model.interfaces.values():

                first = TransitionModelBuilder._first_visible(
                    interface.visible_on,
                    milestones,
                )

                last = TransitionModelBuilder._last_visible(
                    interface.visible_on,
                    milestones,
                )

                if milestone in interface.visible_on:
                    active_interfaces += 1

                if first == milestone:
                    new_interfaces += 1

                if (
                    last == milestone
                    and milestone != final_milestone
                ):
                    retired_interfaces += 1

            sheet.append(
                [
                    milestone,
                    new_nodes,
                    retired_nodes,
                    active_nodes,
                    new_interfaces,
                    retired_interfaces,
                    active_interfaces,
                ]
            )

        ExcelExporter._auto_width(sheet)

    # ---------------------------------------------------------------------

    @staticmethod
    def _write_validation_sheet(
        workbook: Workbook,
        report: ValidationReport,
    ) -> None:

        sheet = workbook.create_sheet("Validation")

        #
        # Summary
        #
        sheet.append(["Metric", "Count"])
        sheet.append(["Errors", report.error_count])
        sheet.append(["Warnings", report.warning_count])
        sheet.append([])

        #
        # Detail header
        #
        headers = [
            "Severity",
            "Rule",
            "Object ID",
            "Object Name",
            "Page",
            "Message",
        ]

        sheet.append(headers)

        #
        # Make the header bold.
        #
        bold = Font(bold=True)

        for cell in sheet[4]:
            cell.font = bold

        #
        # Freeze the header.
        #
        sheet.freeze_panes = "A5"

        #
        # Sort findings:
        # Errors first, then warnings,
        # then by rule, object and page.
        #
        severity_order = {
            ValidationSeverity.ERROR: 0,
            ValidationSeverity.WARNING: 1,
            ValidationSeverity.INFO: 2,
        }

        issues = sorted(
            report.issues,
            key=lambda i: (
                severity_order[i.severity],
                i.rule,
                i.object_id,
                i.page,
            ),
        )

        for issue in issues:

            sheet.append(
                [
                    issue.severity.value,
                    issue.rule,
                    issue.object_id,
                    issue.object_name,
                    issue.page,
                    issue.message,
                ]
            )

        ExcelExporter._auto_width(sheet)
